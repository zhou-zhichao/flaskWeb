import glob
import os
import re
import shutil
import unicodedata  # 导入unicodedata模块

import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook


def common_prefix(folder):
    # 获取文件夹下的所有文件名
    files = os.listdir(folder)
    # 如果文件夹为空，返回空字符串
    if not files:
        return ""
    # 否则，使用min和max函数找到字典序最小和最大的文件名
    min_file = min(files)
    max_file = max(files)
    # 初始化公共前缀为空字符串
    prefix = ""
    # 遍历两个文件名的最短长度
    for i in range(min(len(min_file), len(max_file))):
        # 如果两个文件名在第i个位置相同，将该字符加入公共前缀
        if min_file[i] == max_file[i]:
            prefix += min_file[i]
        # 否则，跳出循环
        else:
            break
    # 返回公共前缀
    return prefix


def copy_sheet(src_xlsx, ssheetname, dst_xlsx, nsheetname=None):
    if nsheetname == None:
        nsheetname = ssheetname
    try:
        sw = load_workbook(f'{src_xlsx}')
    except KeyError:
        raise KeyError('旧工作簿不存在 The old xlsx is not exists')

    try:
        dw = load_workbook(f'{dst_xlsx}')
    except FileNotFoundError:
        dw = Workbook()

    try:
        sheet = dw[f'{nsheetname}']
    except KeyError:
        sheet = dw.create_sheet(f'{nsheetname}')

    try:
        src_sheet = sw[f'{ssheetname}']
    except KeyError:
        raise KeyError('源工作簿文件不存在该工作簿 The sheet does not exist in the source file')

    for row in src_sheet.iter_rows():
        row_list = []
        for cell in row:
            row_list.append(cell.value)
        sheet.append(row_list)
    dw.save(f'{dst_xlsx}')


# 定义一个函数，用于自动调节列宽
def auto_adjust_column_width_writer(writer, df):
    # Get the workbook and worksheet objects
    # workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    # Iterate over the columns and set the column width based on the max length in each column
    for i, col in enumerate(df.columns):
        # find length of column i
        column_len = max(
            df[col].astype(str).apply(
                lambda x: sum(1 + (unicodedata.east_asian_width(c) in "WF") for c in x)
            )
        )
        # Setting the length if the column header is larger than the max column value length
        column_len = max(column_len, sum(1 + (unicodedata.east_asian_width(c) in "WF") for c in col)) + 2
        # set the column length
        worksheet.set_column(i, i, column_len)


def file_auto_adjust_column_width(file_name, sheet_name='Sheet1'):
    # Load the excel file and get the worksheet object
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]
    # Convert the worksheet to a dataframe
    df = pd.DataFrame(worksheet.values)
    # Drop the first row which contains the column headers
    df = df.drop(0, axis=0)
    # Rename the columns with the headers
    df.columns = worksheet[1]
    # Iterate over the columns and set the column width based on the max length in each column
    for i, col in enumerate(df.columns):
        # find length of column i
        column_len = max(
            df[col].astype(str).apply(
                lambda x: sum(1 + (unicodedata.east_asian_width(c) in "WF") for c in x)
            )
        )
        # Setting the length if the column header is larger than the max column value length
        column_len = max(column_len, sum(1 + (unicodedata.east_asian_width(c) in "WF") for c in col.value)) + 2
        # set the column length
        worksheet.column_dimensions[col.column_letter].width = column_len
    # Save and close the workbook
    workbook.save(file_name)
    workbook.close()


def auto_adjust_column_width(df, file_name, sheet_name='Sheet1'):
    # Create a temporary Excel writer object
    writer = pd.ExcelWriter(file_name, engine="xlsxwriter")
    # Write the dataframe to the worksheet
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    # Get the workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    # Iterate over the columns and set the column width based on the max length in each column
    for i, col in enumerate(df.columns):
        # find length of column i
        column_len = max(
            df[col].astype(str).apply(
                lambda x: sum(1 + (unicodedata.east_asian_width(c) in "WF") for c in x)
            )
        )
        # Setting the length if the column header is larger than the max column value length
        column_len = max(column_len, sum(1 + (unicodedata.east_asian_width(c) in "WF") for c in col)) + 2
        # set the column length
        worksheet.set_column(i, i, column_len)
    # Save and close the workbook
    workbook.close()


# 定义一个函数，用于将一个Excel文件中的一个工作表按照"字段确认"列的不同值分割成多个子表，并保存到不同的Excel文件中
def sep_on_field(file_name):
    # Load the data from the specified Excel file and sheet name
    # try:
    df = pd.read_excel(f"{file_name}", engine="openpyxl", sheet_name='外供数据检查')
    df[['业务域', "业务子域", "业务过程", "表/视图名称"]] = df[
        ['业务域', "业务子域", "业务过程", "表/视图名称"]].fillna(
        method='ffill')
    df['字段确认'].fillna('已确认', inplace=True)
    df.dropna(subset=["字段确认"], inplace=True)
    # Get the unique values of the "字段确认"
    distinct_values = df["字段确认"].unique()

    sheets = {}
    # For each unique value, get the subset of the dataframe and store it in the dictionary
    df = df.rename(columns={'所属表': '表/视图名称', '业务过程': '业务单元'})
    for field in distinct_values:
        # sub_df = df[df["字段确认"] == field]
        sub_df = df[df['字段确认'] == field].copy()
        # 这段代码中的警告是因为你在对一个切片的数据框进行赋值操作，这可能会导致不可预期的结果。为了避免这个警告，你可以在切片的时候使用.copy()方法，这样就会创建一个新的数据框对象，而不是一个视图。
        sheets[field] = sub_df
        print(f"Field: {field}, row count: {len(sub_df)}")
    file_name = re.sub(r"\.[^.]+$", "", file_name)

    # Save each subset dataframe to a separate Excel file
    for field, subdiv_df in sheets.items():
        safe_field = field.replace("/", "-").replace(":", "")
        output_dir = "file/temp/" + file_name.rsplit("\\")[-1] + "/"
        # Create an output directory if it does not exist
        # output_dir = file_name
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        # else:
        #     shutil.rmtree(output_dir)
        #     os.makedirs(output_dir)
        # Create an ExcelWriter object with xlsxwriter engine
        final_path = f"{output_dir}{safe_field}.xlsx"
        writer = pd.ExcelWriter(final_path, engine='xlsxwriter')
        # Add two columns to the subset dataframe if the field is '新增'
        if field == '新增':
            subdiv_df.loc[:, '新增类型'] = ''
            subdiv_df.loc[:, '确认状态'] = ''
        # Write the dataframe to the ExcelWriter object
        subdiv_df.to_excel(writer, index=False)
        # Call the auto_adjust_column_width function to adjust the column width
        auto_adjust_column_width_writer(writer, subdiv_df)
        # Save and close the ExcelWriter object
        writer.close()
    return output_dir
    # except FileNotFoundError as e:
    #     print(e)
    #     print("分割文件没有成功\nsep on filed")


def sep_on_sheet(excel):
    # Load the excel file and get the sheet names
    workbook = openpyxl.load_workbook(excel)
    sheet_names = workbook.sheetnames
    # Iterate over the sheet names and create a new excel file for each sheet
    for sheet_name in sheet_names:
        if sheet_name in ["统计结果", "主数据检查"]:
            continue
        # Get the worksheet object for the current sheet
        worksheet = workbook[sheet_name]
        # Create a new workbook and copy the worksheet to it
        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active
        new_worksheet.title = sheet_name
        for row in worksheet:
            for cell in row:
                new_worksheet[cell.coordinate].value = cell.value
        # Save the new workbook with the file name as filname_sheetname.xlsx
        file_name = re.sub(r"\.[^.]+$", "", excel)
        dir_name = file_name.split('/')[-1]
        # dir_name = dir_name.split('_')[0]+dir_name.split('_')[1]
        file_name = dir_name.split("_")[0] + "_" + dir_name.split("_")[1] + '_' + sheet_name + '.xlsx'
        file_path = "file/confirm/" + dir_name
        final_path = os.path.join(file_path, file_name)
        if not os.path.exists(file_path):
            os.makedirs(file_path)
        new_workbook.save(final_path)
        # Call the file_auto_adjust_column_width function to adjust the column width
        file_auto_adjust_column_width(final_path, sheet_name)
    # Close the original workbook
    workbook.close()


def merge_on_field(dir_path, file_name):
    # 创建一个空的列表，用来存放要合并的DataFrame
    df_list = []
    # 遍历文件夹下的所有excel文件
    for file in os.listdir(dir_path):
        # 如果文件是excel格式，就读取它
        if file.endswith(".xlsx"):
            # 读取文件的第一个sheet，忽略索引列
            df = pd.read_excel(os.path.join(dir_path, file), index_col=None)
            # 把文件的数据添加到列表中
            df_list.append(df)
    # 用pd.concat函数把列表中的所有DataFrame合并成一个大的DataFrame
    merged_df = pd.concat(df_list, ignore_index=True)
    if '数据元素uuid' in merged_df.columns:
        merged_df = merged_df.drop(columns='数据元素uuid')
    # 打开file_name的路径的excel文件，如果不存在就创建一个
    # auto_adjust_column_width(merged_df, file_name, "外供数据检查")
    if not os.path.isfile(file_name):
        # 导入openpyxl模块
        # import openpyxl
        # 创建一个新的工作簿对象
        workbook = openpyxl.Workbook()
        # 保存工作簿到文件名
        workbook.save(file_name)
    writer = pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists='replace')
    # 把merged_df写入 外供数据检查 sheet，如果存在就覆盖
    merged_df.to_excel(writer, sheet_name="外供数据检查", index=False)
    # 保存并关闭文件
    # writer.save()
    writer.close()


def merge_to_standard(file_path, merge_file):
    merge_exists = os.path.exists(merge_file)
    for file_name in os.listdir(file_path):
        # If the file name contains '新增' and is an Excel file
        if re.search('新增|特殊', file_name) and file_name.endswith('.xlsx'):
            # Read the Excel file into a dataframe
            df = pd.read_excel(os.path.join(file_path, file_name), engine='openpyxl')
            # Read the merge_file into another dataframe
            if not os.path.exists(merge_file):
                # 如果不存在，就创建一个空的excel文件
                df_empty = pd.DataFrame(
                    columns=["业务域", "业务子域", "业务单元", "表/视图名称", "字段顺序", "字段名", "字段类型", "长度/值域",
                             "精度", "修订建议", "修订确认", "数据元素uuid", "字段确认", "新增类型", "确认状态"])

                writer = pd.ExcelWriter(merge_file, engine="openpyxl")

                df_empty.to_excel(writer, sheet_name="Sheet1", index=False)
                # writer.save()
                writer.close()
            merge_df = pd.read_excel(merge_file, engine='openpyxl')
            # Merge the two dataframes on the columns 所属表 and 字段名, using left join to keep merge_file's values
            merged_df = pd.merge(merge_df, df, on=['表/视图名称', '字段名'], how='left', suffixes=('', '_y'))
            # Drop the duplicate columns from the right dataframe
            merged_df = merged_df.drop(merged_df.filter(regex='_y$').columns.tolist(), axis=1)
            # Fill the NaN values in 新增类型 and 确认状态 columns with the values from df
            # merged_df['新增类型'] = merged_df['新增类型'].fillna(df['新增类型'])
            # merged_df['确认状态'] = merged_df['确认状态'].fillna(df['确认状态'])
            # merged_df['修订建议'] = merged_df['修订建议'].fillna(df['修订建议'])
            # Concatenate the two dataframes along the r axis, keeping only the unique rows
            merged_df = pd.concat([merged_df, df], axis=0).drop_duplicates(keep='last', subset=['表/视图名称', '字段名'])
            if '数据元素uuid' in merged_df.columns:
                merged_df = merged_df.drop(columns='数据元素uuid')
            # Save the merged dataframe to a new Excel file
            auto_adjust_column_width(merged_df, merge_file)
            # merged_df.to_excel(merge_file, index=False)
            # Break the loop
            break


def first_sheet_write(file_path, merge_file):
    # 打开一个excel file_path的'产品线对外数据要求数据检查'sheet
    df = pd.read_excel(file_path, sheet_name='对外数据要求检查')
    # 筛选出标准状态不是'有'的所有行
    df = df[df['标准确认'] != '有']
    df = df[df['业务线确认'] != '暂不提供']
    # 将筛选出的行添加到merge_file的'产品线对外数据要求数据检查'sheet中
    if os.path.exists(merge_file):  # 检查文件是否存在
        print('文件已存在')
    else:  # 如果不存在
        with open(merge_file, 'a') as f:  # 以写入模式打开文件
            print('文件创建成功')
    with pd.ExcelWriter(merge_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='产品线对外数据要求数据检查', index=False)
    # with pd.ExcelWriter(, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    #     df.to_excel(writer, sheet_name='产品线对外数据要求数据检查', index=False)


def data_elem_align(file_path, elem_standard, sheet_name):
    # 读取两个excel文件
    df1 = pd.read_excel(file_path, sheet_name=sheet_name)
    df2 = pd.read_excel(elem_standard)
    # 按照标准表和标准字段两个列来合并两个数据框，保留左边数据框的所有行，用how='left'参数
    df3 = pd.merge(df1, df2, on=['表/视图名称', '字段名'], how='left', suffixes=('', '_y'))
    # 只保留数据元素这一列，用subset参数
    df4 = df3[['数据元素']]
    # 将数据元素这一列添加在file_path后面，用pd.concat函数，axis=1表示按列合并
    if '数据元素' in df1.columns:
        df1 = df1.drop('数据元素', axis=1)
    if '数据元素分类' in df1.columns:
        df1 = df1.drop('数据元素分类', axis=1)
    df5 = pd.concat([df1, df4], axis=1)
    # df5["数据元素"] = df5["数据元素"].replace(df4["数据元素"], df1["数据元素"])
    # 读取一个叫dataelement.xlsx的文件的 数据元素 sheet
    df6 = pd.read_excel('file/dataelement.xlsx', sheet_name='数据元素')
    # 把df5中的数据元素和 dataelement的数据元素名称对应，用pd.merge函数，on='数据元素'表示按照数据元素这一列来合并
    df7 = pd.merge(df5, df6, on='数据元素', how='left')
    # 把df5中的数据元素后面加一列数据元素分类 并把dataelement的数据元素分类添加在 df5的数据元素后面，用subset参数
    df8 = df7[['数据元素分类']]
    # 将数据元素分类这一列添加在df5后面，用pd.concat函数，axis=1表示按列合并
    df9 = pd.concat([df5, df8], axis=1)
    # 返回合并后的数据框
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a',
                        if_sheet_exists='replace') as writer:
        df9.to_excel(writer, sheet_name=sheet_name, index=False)  # todo：多一个sheet
    workbook = openpyxl.load_workbook(file_path)

    # 删除名为sheet的工作表
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    workbook.save(file_path)


def code_check(standard_file, append_file):
    # 读取两个excel文件
    df1 = pd.read_excel(standard_file)
    df2 = pd.read_excel(append_file)
    # 按照标准表和标准字段两个列来合并两个数据框，保留左边数据框的所有行，用how='left'参数
    df2["代码"] = df2["代码"].astype('object')
    df3 = pd.merge(df1, df2, on=['代码名称', '代码'], how='right', suffixes=('_y', ''))
    # df3 = df3[['代码名称'	,'代码编码',	'代码',	'上级编码']]
    selected_columns = ['代码名称', '代码编码', '代码', '上级编码', '代码编码_y', '上级编码_y']
    df = df3[selected_columns]
    df['上级编码'] = df['上级编码'].apply(lambda x: str(x).split('.')[0])
    # df['上级编码_y'] = df['上级编码_y'].apply(lambda x: str(x).split('.')[0])
    df['代码编码'] = df['代码编码'].astype(str)
    df['上级编码'] = df['上级编码'].replace(np.nan, 'nan')
    df['上级编码_y'] = df['上级编码_y'].replace(np.nan, 'nan')
    df['代码编码_y'] = df['代码编码_y'].replace(np.nan, 'nan')

    # 创建一个条件列表
    conditions = [
        (df['代码编码_y'] == 'nan') & (df['上级编码_y'] == 'nan'),  # 代码编码_y和上级编码_y都为 nan
        (df['代码编码'] == df['代码编码_y']) & (df['上级编码'] == df['上级编码_y']),  # 代码编码和上级编码都一致
        (df['代码编码'] != df['代码编码_y']) & (df['上级编码'] == df['上级编码_y']),  # 代码编码不一致，上级编码一致
        (df['代码编码'] == df['代码编码_y']) & (df['上级编码'] != df['上级编码_y']),  # 代码编码一致，上级编码不一致
        (df['代码编码'] != df['代码编码_y']) & (df['上级编码'] != df['上级编码_y'])  # 代码编码和上级编码都不一致

    ]
    # 创建一个值列表
    values = [
        '新增',  # 代码编码_y和上级编码_y都为 nan
        '正常',  # 代码编码和上级编码都一致
        '代码编码不一致',  # 代码编码不一致，上级编码一致
        '上级编码不一致',  # 代码编码一致，上级编码不一致
        '代码和上级都不一致',  # 代码编码和上级编码都不一致

    ]
    # 使用 np.select 方法创建新列
    df['状态'] = np.select(conditions, values)
    # df = df[df['状态'] != '正常']

    df.columns = ['代码名称', '代码编码', '代码', '上级编码', '代码编码_标准', '上级编码_标准', '状态']
    file_name = append_file.split('/')[-1]
    dir_name = 'file/confirm/' + file_name.rsplit("_", 1)[0] + "_重要业务结果/"
    check_and_create_folder(dir_name)
    auto_adjust_column_width(df, dir_name + file_name)
    print("别紧张，只是警告。功能正常运行。")


def statistics(ana_file):
    master_data, standard_output, business_demand = pd.read_excel(ana_file,
                                                                  sheet_name=['主数据检查', '重要业务结果检查',
                                                                              '外部数据检查']).values()
    master_data['是否提供'] = master_data['是否提供'] == '提供'
    result = master_data.groupby('主数据').agg({'主数据': 'nunique', '是否提供': 'any'})
    total = result.sum()
    first = total['主数据']
    second = total['是否提供']
    standard_master_num = master_data['主数据'].count()  # 标准主数据属性数

    business_master_num = master_data.是否提供.value_counts().loc[True]  # 业务线主数据属性数

    standard_business_unit_num_list = pd.concat(
        [standard_output['业务域'], standard_output['业务子域'], standard_output['业务单元']],
        axis=1).drop_duplicates().count()
    standard_business_unit_num = standard_business_unit_num_list['业务域']
    drop_standard_output = standard_output[standard_output.字段确认 != '业务线缺']
    app_num_list = pd.concat(
        [drop_standard_output['业务域'], drop_standard_output['业务子域'], drop_standard_output['业务单元']],
        axis=1).drop_duplicates().count()
    app_num = app_num_list['业务域']
    standard_output_num = standard_output['业务域'].count()
    drop_drop_standard_output = drop_standard_output[drop_standard_output.字段确认 != '新增']
    business_output_num = drop_standard_output['业务域'].count()
    elem_num = drop_standard_output['数据元素'].count()
    overlapping_num = drop_drop_standard_output['业务域'].count()
    drop_business_demand = business_demand[business_demand.确认状态 != '删除']
    drop_business_demand = drop_business_demand[drop_business_demand.确认状态 != '删除，有多个收录']
    demand_num = drop_business_demand.类型.count()
    in_standard_num = drop_business_demand['标准状态'].value_counts().loc['有'] + drop_business_demand[
        '确认状态'].str.contains('增加').sum()
    master_cover_ratio = business_master_num / standard_master_num
    business_cover_ratio = app_num / standard_business_unit_num
    implementation_ratio = overlapping_num / standard_output_num
    out_demand_cover_ratio = in_standard_num / demand_num

    import shutil
    shutil.copy("standard/统计模板.xlsx", 'output/统计.xlsx')
    from openpyxl import load_workbook
    # 打开一个名为data.xlsx的文件
    workbook = load_workbook(filename='output/统计.xlsx')
    # 获取第一个工作表
    sheet = workbook.active

    positions = ['B2', 'B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B9', 'B10', 'B11', 'B12', 'B13', 'B14', 'B15', 'B16', 'B17',
                 'B18', 'B19', 'B20', 'B21', 'B22', 'E25', 'E26', 'E27', 'E28']
    # positions = ['B2', 'B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B9', 'B10', 'B11', 'B12','B13']
    vals = [first, second, standard_master_num, business_master_num, standard_business_unit_num, app_num,
            standard_business_unit_num - app_num, overlapping_num, standard_output_num, business_output_num, '', '',
            demand_num, in_standard_num,
            master_cover_ratio, business_cover_ratio, implementation_ratio, business_output_num - elem_num, elem_num,
            elem_num / business_output_num, out_demand_cover_ratio,
            master_cover_ratio,
            business_cover_ratio, implementation_ratio, out_demand_cover_ratio]
    # print(len(positions), len(vals))
    for position, val in zip(positions, vals):
        sheet[position].value = val
    # worksheet.write('B5', 100)
    sheet_inner = workbook.create_sheet('重要业务结果检查')
    sheet_outer = workbook.create_sheet('外部数据检查')
    temp_workbook = openpyxl.load_workbook('input/人事业务线数据标准落标分析.xlsx')
    source_inner = temp_workbook['重要业务结果检查']
    source_outer = temp_workbook['外部数据检查']
    for row in source_inner.iter_rows(values_only=True):
        sheet_inner.append(row)
    for row in source_outer.iter_rows(values_only=True):
        sheet_outer.append(row)
    # workbook.move_sheet('统计结果', after='外部数据检查')
    # workbook.reorder_sheets(['重要业务结果检查', '外部数据检查', '统计结果'])
    workbook.save('output/统计.xlsx')

    # workbook.save('output/统计.xlsx')
    # file_auto_adjust_column_width('output/统计.xlsx')
    # copy_sheet('output/统计.xlsx', '统计结果', 'output/test.xlsx')
    # workbook = load_workbook(filename='output/统计.xlsx')
    # sheet = workbook['统计结果']
    # target = workbook.copy_worksheet(sheet)
    # workbook.save(filename="wb.xlsx")
    pass


def elem_align(file_path, elem_standard):
    # 读取两个excel文件
    df1 = pd.read_excel(file_path, sheet_name='重要业务结果检查')
    df2 = pd.read_excel(elem_standard)
    # 按照标准表和标准字段两个列来合并两个数据框，保留左边数据框的所有行，用how='left'参数
    df3 = pd.merge(df1, df2, on=['标准表', '标准字段'], how='left')
    # 只保留数据元素这一列，用subset参数
    df4 = df3[['数据元素']]
    # 将数据元素这一列添加在file_path后面，用pd.concat函数，axis=1表示按列合并
    df5 = pd.concat([df1, df4], axis=1)
    # 读取一个叫dataelement.xlsx的文件的 数据元素 sheet
    df6 = pd.read_excel('standard/dataelement.xlsx', sheet_name='数据元素')
    # 把df5中的数据元素和 dataelement的数据元素名称对应，用pd.merge函数，on='数据元素'表示按照数据元素这一列来合并
    df7 = pd.merge(df5, df6, on='数据元素', how='left')
    # 把df5中的数据元素后面加一列数据元素分类 并把dataelement的数据元素分类添加在 df5的数据元素后面，用subset参数
    df8 = df7[['数据元素分类']]
    # 将数据元素分类这一列添加在df5后面，用pd.concat函数，axis=1表示按列合并
    df9 = pd.concat([df5, df8], axis=1)
    # 返回合并后的数据框
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a',
                        if_sheet_exists='replace') as writer:
        df9.to_excel(writer, sheet_name='重要业务结果检查', index=False)


def delete_all(folder):
    # 遍历文件夹下的所有文件和子文件夹
    for file in glob.glob(folder + "*"):
        # 如果是文件，就删除
        if os.path.isfile(file):
            os.remove(file)
        # 如果是文件夹，就递归调用自己
        elif os.path.isdir(file):
            delete_all(file + "/")
            # 删除空文件夹
            os.rmdir(file)


def confirm_list_to_tuple(file_list):
    # 创建一个空列表来存储元组
    tuple_list = []
    # 创建一个空字典来存储三层文件名和对应的四层文件名
    file_name = {}
    # 遍历文件列表中的每个文件名
    for i, file in enumerate(file_list):
        all_part = file.split("\\")
        part_len = len(all_part)
        if part_len == 4:
            # 用反斜杠分割文件名，取最后两个部分
            parts = file.split("\\")[-2:]
            # 用下划线分割第一个部分，取前两个部分
            subparts = parts[0].split("_")[:2]
            # 将第二个部分添加到子部分列表中
            subparts.append(parts[1])
            file_name[parts[1]] = i
            # 将子部分列表转换为元组，并添加到元组列表中
            subparts.append(file)
            tuple_list.append(subparts)


        elif part_len == 3:
            # 用反斜杠分割文件名，取最后一个部分
            part = file.split("\\")[-1]
            name = part.rsplit(".", 1)[0].replace("确认", "") + ".xlsx"
            # 在字典中查找是否有对应的四层文件名，如果有则添加到子部分列表中，否则添加空字符串
            if name in file_name:
                add_part = tuple_list[file_name[name]]
                add_part.append(part)
                add_part.append(file)
                # subparts.append("file\\confirm\\" + part + "\\" + file_name[part])
            else:
                add_part.append("")
                add_part.append("")
            # 将三层文件名添加到子部分列表中

    # 返回元组列表
    return tuple_list


def transform_list(file_list):
    # 定义一个空的字典
    dic = {}
    # 遍历列表中的每个文件路径
    for path in file_list:
        # 用反斜杠分割成子字符串
        sub = path.split('\\')
        # 取出最后两个子字符串
        folder = sub[-2]
        file = sub[-1]
        # 判断字典中是否已经有这个文件夹名作为键
        if folder not in dic:
            # 创建一个新的键值对
            dic[folder] = [file]
        else:
            # 把文件名追加到对应的列表中
            dic[folder].append(file)
    # 定义一个空的列表
    result = []
    # 遍历字典中的每个键值对
    for key, value in dic.items():
        # 用文件夹名拆分成前两位和后一位
        first, second = key.split('_')[:2]
        third = ''
        fourth = ''
        # 遍历文件名列表中的元素
        for name in value:
            # 判断是否有重要业务结果或自定义代码
            if '重要业务结果' in name:
                third = name
            elif '自定义代码' in name:
                fourth = name

        standard_elem_name = first + "_业务线数据元素映射.xlsx"
        elem_exists = os.path.exists(os.path.join("file/standard/", standard_elem_name))
        if elem_exists:
            fifth = standard_elem_name
        else:
            fifth = ""
        # 把每个元组添加到结果列表中
        result.append((first, second, third, fourth, fifth))
    # 返回结果列表
    return result


def get_all_filenames(path):
    # 创建一个空列表，用来存储所有文件名
    filenames = []
    # 遍历指定目录及其子目录
    for root, dirs, files in os.walk(path):
        # 对于每个非目录子文件
        for file in files:
            # 获取文件的完整路径
            file_path = os.path.join(root, file)
            # 把文件名添加到列表中
            filenames.append(file_path)
    # 返回文件名列表
    return filenames


def check_and_create_folder(folder_path):
    # 检查文件夹路径是否有效
    if not isinstance(folder_path, str):
        print("无效的文件夹路径")
        return
    # 检查文件夹是否存在
    if os.path.exists(folder_path):
        print("文件夹已存在")
    else:
        # 创建文件夹
        try:
            os.makedirs(folder_path)
            print("文件夹创建成功")
        except OSError as e:
            print("文件夹创建失败，错误信息：", e)


def read_filenames_with_version(folder):
    # 创建一个空列表来存储结果
    result = []
    # 遍历文件夹下的所有文件
    for file in os.listdir(folder):
        # 把文件名按照 _ 拆分成三段
        segments = file.split("_")
        # 如果拆分后的长度不是3，说明文件名不符合要求，跳过这个文件
        if len(segments) != 3:
            continue
        # 否则，把拆分后的三段分别赋值给业务线、版本、文件类型
        business, version, file_type = segments
        # 把这三个值作为一个元组添加到结果列表中
        result.append((business, version, file_type, file))
    # 返回结果列表
    return result


def read_filenames(folder):
    # 创建一个空列表来存储结果
    result = []
    # 遍历文件夹下的所有文件
    for file in os.listdir(folder):
        # 把文件名按照 _ 拆分成2段
        segments = file.split("_")
        # 如果拆分后的长度不是3，说明文件名不符合要求，跳过这个文件
        if len(segments) != 2:
            file_type = file.split(".")[0]
            business = ""
        else:
            # 否则，把拆分后的三段分别赋值给业务线、版本、文件类型
            business, file_type = segments
            file_type = file_type.split('.')[0]
        # 把这三个值作为一个元组添加到结果列表中
        result.append((business, file_type, file))
    # 返回结果列表
    return result


def xlsx_func(filename):
    field = filename.rsplit("_")[-1].split(".")[0]
    if field == "外供数据检查确认":
        dirname = sep_on_field(filename)
        version = filename.rsplit('\\')[-1].rsplit("_", 1)[0]
        check_and_create_folder("file/modify/")
        merge_to_standard(dirname, f"file/modify/{version}_重要结果修订.xlsx")
        merge_on_field(dirname, f"file/temp/{version}_外供数据检查确认.xlsx")
        data_elem_align(f"file/temp/{version}_外供数据检查确认.xlsx", "file/standard/" +
                        version.split("_")[0] + "_业务线数据元素映射.xlsx", "外供数据检查")
        source_path = "file/input/" + version +f"{version}_重要业务结果.xlsx"
        shutil.copy(source_path, f"file/temp/{version}_重要业务结果.xlsx")
        r = openpyxl.load_workbook("r.xlsx")
        w = openpyxl.load_workbook("w.xlsx")

        # 获取两个文件中的所有工作表的名字
        r_sheets = r.sheetnames
        w_sheets = w.sheetnames

        # 遍历r文件中的每一个工作表
        for r_sheet in r_sheets:
            # 判断是否在w文件中存在同名的工作表
            if r_sheet in w_sheets:
                # 获取r文件中对应工作表的数据
                r_data = r[r_sheet].values
                # 获取w文件中对应工作表的最大行数和最大列数
                max_row = w[r_sheet].max_row
                max_col = w[r_sheet].max_column
                # 遍历r_data中的每一行数据
                for i, row in enumerate(r_data):
                    # 遍历每一列数据
                    for j, value in enumerate(row):
                        # 将数据写入w文件中对应工作表的相同位置
                        w[r_sheet].cell(row=i + 1, column=j + 1, value=value)

        # 保存w文件，并关闭两个文件
        w.save("w.xlsx")
        r.close()
        w.close()
        statistics(f"file/temp/{version}_外供数据检查确认.xlsx")
    elif field == "对外数据要求检查确认":
        pass
    elif field == "自定义代码确认":
        pass
    else:
        print("文件名出错")
